from __future__ import annotations

import argparse
import hashlib
import json
import re
import unicodedata
from pathlib import Path
from typing import Any
from datetime import date

from openpyxl import load_workbook

SYSTEM_ID = "lo3rwang"
PRIMARY_LOC = "LOC3"


def shared_registry_path(name: str) -> Path:
    return Path(__file__).resolve().parents[2] / "data" / "shared" / name


def load_loc3_author_annotations() -> dict[str, dict[str, Any]]:
    path = shared_registry_path("LOC3_AUTHOR_ANNOTATION_REGISTRY.json")
    if not path.exists():
        return {}
    payload = json.loads(path.read_text(encoding="utf-8"))
    return {
        text(item.get("work_id")): item
        for item in payload.get("records", [])
        if text(item.get("work_id"))
    }


def load_era_registry() -> tuple[list[dict[str, Any]], dict[str, dict[str, Any]], dict[str, Any] | None]:
    path = shared_registry_path("LOC_ERA_REGISTRY.json")
    if not path.exists():
        return [], {}, None
    payload = json.loads(path.read_text(encoding="utf-8"))
    eras = payload.get("eras", [])
    by_period = {text(item.get("period")): item for item in eras if text(item.get("period"))}
    current = next((item for item in eras if text(item.get("status")) == "current"), None)
    return eras, by_period, current


def parse_iso_date(value: Any) -> date | None:
    raw = text(value)
    if not raw:
        return None
    try:
        return date.fromisoformat(raw[:10])
    except ValueError:
        return None


def resolve_era(created_date: Any, period: str, eras: list[dict[str, Any]], by_period: dict[str, dict[str, Any]]) -> dict[str, Any]:
    """Resolve ERA from the work date first; period is only a fallback."""
    d = parse_iso_date(created_date)
    if d:
        for item in eras:
            start = parse_iso_date(item.get("start_date"))
            end = parse_iso_date(item.get("end_date"))
            if not start:
                continue
            if d >= start and (end is None or d <= end):
                return item
    return by_period.get(period, {})


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def number(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def split_values(value: Any) -> list[str]:
    return [part.strip() for part in re.split(r"[｜|;\n]+", text(value)) if part.strip()]


def first_nonempty(row: dict[str, Any], *fields: str) -> str:
    """Return the first non-empty source field without inventing missing provenance."""
    for field in fields:
        value = text(row.get(field))
        if value:
            return value
    return ""


def build_rune_provenance(row: dict[str, Any]) -> dict[str, Any] | None:
    """Preserve rune/draw creation provenance when the source workbook actually records it."""
    origin = first_nonempty(
        row,
        "創作來源",
        "作品來源",
        "歌曲來源",
        "生成來源",
        "靈感來源",
        "創作方式",
    )
    rune_song_flag = first_nonempty(
        row,
        "符文歌曲",
        "是否符文歌曲",
        "抽卡歌曲",
        "是否抽卡歌曲",
    )
    draw_mode = first_nonempty(
        row,
        "抽牌模式",
        "抽卡模式",
        "牌陣",
        "Spread",
    )
    draw_date = first_nonempty(
        row,
        "抽牌日期",
        "抽卡日期",
        "占卜日期",
    )
    draw_result = first_nonempty(
        row,
        "抽牌結果",
        "抽卡結果",
        "符文結果",
        "符文組合",
        "OW3gs結果",
    )
    source_note = first_nonempty(
        row,
        "創作註記",
        "來源註記",
        "備註",
        "Notes",
    )

    values = {
        "source_origin": origin or None,
        "rune_song_flag": rune_song_flag or None,
        "draw_mode": draw_mode or None,
        "draw_date": draw_date or None,
        "draw_result": draw_result or None,
        "source_note": source_note or None,
    }
    if not any(values.values()):
        return None
    return values


LOC3_EXPLICIT_EXCLUDED_WORK_IDS = {
    "E0216",  # 啾咪十八歲 — author-confirmed playful/experimental exception
}


def is_loc3_analysis_exception(row: dict[str, Any]) -> bool:
    """Exclude author-confirmed experiments from formal LOC3 analysis/search."""
    work_id = text(row.get("作品ID"))
    title = text(row.get("代表歌名"))
    if work_id in LOC3_EXPLICIT_EXCLUDED_WORK_IDS:
        return True
    # Author governance rule: titles longer than 16 characters are experimental/problematic
    # generation attempts and are not part of the formal analyzable LOC3 corpus.
    if len(title) > 16:
        return True
    return False


def normalized_lyrics(value: Any) -> str:
    value = unicodedata.normalize("NFKC", text(value)).lower()
    value = re.sub(r"\[[^\]]*\]|【[^】]*】", "", value)
    # Some historical lyrics were saved with spaces between CJK characters.
    # Remove those spaces before semantic/keyword normalization to avoid false misses.
    value = re.sub(r"(?<=[\u3400-\u9fff])\s+(?=[\u3400-\u9fff])", "", value)
    value = re.sub(r"[^\w\u3400-\u9fff]+", "", value, flags=re.UNICODE)
    return value


def sheet_rows(workbook, name: str) -> list[dict[str, Any]]:
    rows = workbook[name].iter_rows(values_only=True)
    headers = [text(value) for value in next(rows)]
    return [dict(zip(headers, row)) for row in rows if any(value is not None for value in row)]


def version_score(version: dict[str, Any], preference: dict[str, Any]) -> float:
    score = number(preference.get("作者版本分數")) * 100
    score += number(preference.get("人聲品質分數")) * 12
    score += number(preference.get("編曲品質分數")) * 10
    score += min(number(version.get("Suno播放次數")), 10000) / 100
    score += min(number(version.get("Suno按讚數")), 1000) / 20
    if text(preference.get("版本推薦狀態")) in {"推薦", "主推", "是"}:
        score += 60
    if text(version.get("是否代表版本")) == "是":
        score += 20
    # Cross-media completion boosts recommendation priority without altering lyric semantics.
    # Reels is a lighter preview asset; a full YouTube MV carries a larger completion bonus.
    if text(version.get("IG短片網址")):
        score += 20
    if text(version.get("YouTube_MV網址")):
        score += 30
    return score


def public_version(version: dict[str, Any], preference: dict[str, Any], work_bonus: float = 0.0) -> dict[str, Any]:
    return {
        "song_id": text(version.get("歌曲ID")),
        "title": text(version.get("歌名")),
        "suno_url": text(version.get("Suno網址")),
        "plays": int(number(version.get("Suno播放次數"))),
        "likes": int(number(version.get("Suno按讚數"))),
        "style_prompt": text(version.get("曲風提示")),
        "playlists": split_values(version.get("播放清單")),
        "ig_preview_url": text(version.get("IG短片網址")),
        "youtube_mv_url": text(version.get("YouTube_MV網址")),
        "author_score": number(preference.get("作者版本分數")),
        "recommendation": text(preference.get("版本推薦狀態")),
        "selection_score": round(version_score(version, preference) + work_bonus, 3),
    }


def resolve_lyric_type(row: dict[str, Any], annotation: dict[str, Any]) -> str:
    explicit = text(annotation.get("lyric_category"))
    if explicit:
        return explicit

    mode = text(annotation.get("discourse_mode"))
    if mode == "state_description":
        return "狀態紀錄型"
    if mode == "narrative_scene":
        return "敘事角色型"
    if mode in {"rational_discourse", "rational_reflection", "system_observation"}:
        return "思辨論述型"

    loc4 = text(row.get("LOC4關聯"))
    note = text(annotation.get("author_note"))
    evidence = f"{loc4} {note}"
    if loc4 and any(token in evidence for token in ("主題曲", "角色", "OP", "特色曲", "小說", "開場", "求婚歌")):
        return "敘事角色型"

    source_category = text(row.get("歌曲主類別"))
    if source_category == "治理宣言型":
        return "治理宣言型"

    return "感性情緒型"


def build(source: Path) -> dict[str, Any]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    works = sheet_rows(workbook, "500公開作品主表")
    versions = sheet_rows(workbook, "663公開版本")
    preferences = sheet_rows(workbook, "版本偏好")
    eras, era_by_period, active_era = load_era_registry()
    author_annotations = load_loc3_author_annotations()

    versions_by_work: dict[str, list[dict[str, Any]]] = {}
    for version in versions:
        versions_by_work.setdefault(text(version.get("作品ID")), []).append(version)
    preference_by_song = {
        text(row.get("歌曲ID")): row for row in preferences if text(row.get("歌曲ID"))
    }

    groups: dict[str, list[dict[str, Any]]] = {}
    for row in works:
        work_id = text(row.get("作品ID"))
        annotation = author_annotations.get(work_id, {})
        search_flag = first_nonempty(row, "向量索引資格", "搜尋資格")
        display_policy = text(row.get("展示政策"))
        if (
            search_flag != "是"
            or display_policy.startswith("hidden")
            or is_loc3_analysis_exception(row)
            or text(annotation.get("analysis_policy")) == "exception_excluded"
        ):
            continue
        lyrics_key = normalized_lyrics(row.get("歌詞"))
        if not lyrics_key:
            continue
        digest = hashlib.sha256(lyrics_key.encode("utf-8")).hexdigest()
        groups.setdefault(digest, []).append(row)

    output_works = []
    for index, (digest, rows) in enumerate(sorted(groups.items()), start=1):
        representative = max(
            rows,
            key=lambda row: (
                number(row.get("最高單曲播放數")),
                number(row.get("全版本播放總數")),
                text(row.get("作品ID")),
            ),
        )
        work_bonus = number(representative.get("推薦加權"))
        loc4_relation = text(representative.get("LOC4關聯"))
        merged_versions = []
        seen_songs = set()
        for row in rows:
            for version in versions_by_work.get(text(row.get("作品ID")), []):
                song_id = text(version.get("歌曲ID"))
                if not song_id or song_id in seen_songs:
                    continue
                seen_songs.add(song_id)
                merged_versions.append(public_version(version, preference_by_song.get(song_id, {}), work_bonus))
        merged_versions.sort(key=lambda item: (-item["selection_score"], -item["plays"], item["song_id"]))

        category = text(representative.get("自動建議主類別")) or text(representative.get("歌曲主類別"))
        start = text(representative.get("自動建議起始狀態")) or text(representative.get("起始狀態"))
        turn = text(representative.get("自動建議轉折方式")) or text(representative.get("轉折方式"))
        final = text(representative.get("自動建議最終狀態")) or text(representative.get("最終狀態"))
        emotion_function = text(representative.get("自動建議情緒功能")) or text(representative.get("情緒功能"))
        tags = []
        for field in ("主題標籤", "情緒標籤", "意象標籤", "情境標籤"):
            tags.extend(split_values(representative.get(field)))
        tags = list(dict.fromkeys(tags))
        annotation = author_annotations.get(text(representative.get("作品ID")), {})
        reasoning_tags = list(dict.fromkeys(annotation.get("reasoning_tags", [])))
        semantic_keywords = list(dict.fromkeys([*reasoning_tags, *tags]))

        retrieval_parts = [
            text(representative.get("代表歌名")), text(representative.get("所有歌名")),
            text(representative.get("AI摘要")), category, start, turn, final,
            emotion_function, text(representative.get("希望延伸")),
            text(representative.get("結尾結構")), text(representative.get("留白意圖")),
            *reasoning_tags,
            text(annotation.get("author_note")),
            *tags,
        ]
        period = text(representative.get("統計時期代碼"))
        era = resolve_era(representative.get("建立日期"), period, eras, era_by_period)
        output_works.append({
            "system_id": SYSTEM_ID,
            "primary_loc": PRIMARY_LOC,
            "related_locs": list(dict.fromkeys(
                ["LOC6", "LOC7", "LOC8"]
                + (["LOC4"] if loc4_relation else [])
                + split_values(representative.get("related_locs"))
            )),
            "work_id": f"LW{index:04d}",
            "lyrics_hash": digest,
            "source_work_ids": [text(row.get("作品ID")) for row in rows],
            "title": text(representative.get("代表歌名")),
            "created_date": text(representative.get("建立日期")),
            "period": period,
            "era_id": text(era.get("era_id")) or None,
            "era": text(era.get("era_id")) or text(representative.get("ERA代碼")),
            "era_name": text(era.get("name")) or text(representative.get("ERA名稱")),
            "playlists": split_values(representative.get("播放清單")),
            "style": text(representative.get("曲風分類")),
            "summary": text(representative.get("AI摘要")),
            "category": category,
            "lyric_type": resolve_lyric_type(representative, annotation),
            "start_state": start,
            "turn_method": turn,
            "final_state": final,
            "emotion_function": emotion_function,
            "discourse_mode": text(annotation.get("discourse_mode")) or None,
            "emotion_applicability": text(annotation.get("emotion_applicability")) or "required",
            "reasoning_tags": reasoning_tags,
            "semantic_keywords": semantic_keywords,
            "author_semantic_note": text(annotation.get("author_note")) or None,
            "semantic_completion": (
                "complete_without_emotion"
                if text(annotation.get("emotion_applicability")) in {"not_primary", "not_required"}
                else "standard"
            ),
            "ending_structure": text(representative.get("結尾結構")),
            "hope_extension": text(representative.get("希望延伸")),
            "tags": tags,
            "content_origin": text(representative.get("content_origin")) or None,
            "loc4_relation": loc4_relation or None,
            "display_policy": text(representative.get("展示政策")) or "normal",
            "recommendation_bonus": work_bonus,
            "rune_provenance": build_rune_provenance(representative),
            "retrieval_text": "\n".join(part for part in retrieval_parts if part),
            "versions": merged_versions,
        })

    return {
        "dataset": {
            "name": "LOC3 Lyrics Search",
            "version": "0.2.0",
            "source": source.name,
            "language_scope": "zh-Hant",
            "unit": "unique_lyrics_work",
            "excluded": ["P1", "non_zh_Hant", "script_pending", "not_searchable", "title_length_gt_16_experiment", "explicit_playful_exceptions"],
            "work_count": len(output_works),
            "version_count": sum(len(item["versions"]) for item in output_works),
            "system_id": SYSTEM_ID,
            "primary_loc": PRIMARY_LOC,
            "era_registry": "data/shared/LOC_ERA_REGISTRY.json",
            "active_era": active_era,
        },
        "works": output_works,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Build the deployable LOC3 lyrics search dataset.")
    parser.add_argument("source", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()
    payload = build(args.source)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    works = payload.pop("works")
    middle = (len(works) + 1) // 2
    shard_names = []
    for number, shard_works in enumerate((works[:middle], works[middle:]), start=1):
        shard_name = f"{args.output.stem}.part{number}.json"
        shard_names.append(shard_name)
        (args.output.parent / shard_name).write_text(
            json.dumps({"works": shard_works}, ensure_ascii=False, separators=(",", ":")),
            encoding="utf-8",
        )
    payload["shards"] = shard_names
    args.output.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(json.dumps(payload["dataset"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
